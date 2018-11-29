# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook

class EmptyTitleRemoverPipeline(object):

    def process_item(self, item, spider):
        if item['job_title']:
            return item


class XLSXWriterPipeline(object):
    """Save scraped data to an XLSX spreadsheet"""

    def open_spider(self, spider):
        """Start scraping"""
        # Create an Excel workbook
        self._wb = Workbook()
        try:
            self._wb = load_workbook('jobs.xlsx')
            self._ws = self._wb.active
        except:
            # Select the active spreadsheet
            self._ws = self._wb.active
            self._ws.title = 'Jobs'
            self._ws.append(['Title', 'Company', 'Crawled Date', 'Posted Date', 'Job Desc', 'Job URL'])
            row = list(self._ws.rows)[0]
            for cell in row:
                cell.font = Font(bold=True)



    def process_item(self, item, spider):
        # Append a row to the spreadsheet
        desc = ''.join(item['job_description'])
        self._ws.append([
            item['job_title'],
            item['company_name'],
            item['crawled_date'],
            item['posted_date'],
            desc,
            item['job_url']
        ])
        return item

    def close_spider(self, spider):
        """Stop scraping"""
        # Save the Excel workbook
        self._wb.save('jobs.xlsx')
